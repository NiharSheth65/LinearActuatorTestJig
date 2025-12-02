import customtkinter as ctk
from tkinter import messagebox
from serial.tools import list_ports
import serial
import time
import win32print
import win32ui
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook



# ---- Initialize theme ----
ctk.set_appearance_mode("DARK")
ctk.set_default_color_theme("blue")

def find_arduino():
    """Automatically detect Arduino COM port."""
    ports = list_ports.comports()
    for port in ports:
        # Filter by common identifiers for Arduino boards
        if "Arduino" in port.description or "CH340" in port.description:
            return port.device  # e.g., 'COM3' or '/dev/ttyUSB0'
    return None

def connect_arduino(retry=True):
    """Try to connect to Arduino dynamically."""
    arduino_port = find_arduino()
    if arduino_port:
        try:
            arduino = serial.Serial(arduino_port, 9600, timeout=0.05)
            time.sleep(2)  # allow Arduino to reset
            print(f"[INFO] Connected to Arduino on {arduino_port}")
            return arduino
        except:
            print(f"[WARN] Failed to connect on {arduino_port}")
            return None
    else:
        print("[WARN] Arduino not found!")
        if retry:
            # Retry until Arduino is plugged in
            while True:
                time.sleep(2)
                arduino = connect_arduino(retry=False)
                if arduino:
                    return arduino
        return None

def send_command(cmd):
    if arduino:
        arduino.write(f"{cmd}\n".encode())
        print(f"[Python] Sent: {cmd}")



class ActuatorGUI:
    def __init__(self, root):
        self.root = root
        root.title("Linear Actuator Jig Test")
        root.geometry("500x520")
        root.resizable(True, True)

        # ---- Header ----
        self.header = ctk.CTkLabel(root, text="Linear Actuator Jig", font=ctk.CTkFont(size=24, weight="bold"))
        self.header.pack(pady=15)

        # ---- Actuator ID ----
        self.id_frame = ctk.CTkFrame(root)
        self.id_frame.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(self.id_frame, text="Actuator ID / Serial:", font=ctk.CTkFont(size=14)).pack(pady=5, anchor="w")
        self.actuator_id_entry = ctk.CTkEntry(self.id_frame, font=ctk.CTkFont(size=14))
        self.actuator_id_entry.pack(pady=5, fill="x")

        # ---- Actuator Type ----
        self.type_frame = ctk.CTkFrame(root)
        self.type_frame.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(self.type_frame, text="Select Actuator Type:", font=ctk.CTkFont(size=14)).pack(anchor="w", pady=5)

        self.act_type = ctk.StringVar(value="100mm")
        self.radio_100 = ctk.CTkRadioButton(self.type_frame, text="100 mm", variable=self.act_type, value="100mm")
        self.radio_100.pack(anchor="w", padx=10)
        self.radio_50 = ctk.CTkRadioButton(self.type_frame, text="50 mm", variable=self.act_type, value="50mm")
        self.radio_50.pack(anchor="w", padx=10)

        # ---- Force Display ----
        self.force_frame = ctk.CTkFrame(root)
        self.force_frame.pack(pady=15, padx=20, fill="x")

        # ctk.CTkLabel(self.force_frame, text="Push Force:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        # self.push_force_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        # self.push_force_label.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        # ctk.CTkLabel(self.force_frame, text="Backdrive Force:", font=ctk.CTkFont(size=14)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        # self.backdrive_force_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        # self.backdrive_force_label.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        
        # Push Force
        ctk.CTkLabel(self.force_frame, text="Push Force:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.push_force_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        self.push_force_label.grid(row=0, column=1, padx=30, pady=5, sticky="w")

        # Push Test Status
        ctk.CTkLabel(self.force_frame, text="Push Test Status:", font=ctk.CTkFont(size=14)).grid(row=0, column=2, padx=10, pady=5, sticky="w")
        self.push_test_status_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        self.push_test_status_label.grid(row=0, column=3, padx=30, pady=5, sticky="w")

        # Backdrive Force
        ctk.CTkLabel(self.force_frame, text="Backdrive Force:", font=ctk.CTkFont(size=14)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.backdrive_force_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        self.backdrive_force_label.grid(row=1, column=1, padx=30, pady=5, sticky="w")

        # Backdrive Test Status
        ctk.CTkLabel(self.force_frame, text="Backdrive Test Status:", font=ctk.CTkFont(size=14)).grid(row=1, column=2, padx=10, pady=5, sticky="w")
        self.backdrive_test_status_label = ctk.CTkLabel(self.force_frame, text="---", font=ctk.CTkFont(size=14), text_color="blue")
        self.backdrive_test_status_label.grid(row=1, column=3, padx=30, pady=5, sticky="w")

        # ---- Control Buttons ----
        self.btn_frame = ctk.CTkFrame(root)
        self.btn_frame.pack(pady=10, padx=20)

        # Emergency Stop
        self.em_stop_btn = ctk.CTkButton(self.btn_frame, text="EMERGENCY STOP", fg_color="red",
                                         hover_color="#ff5555", width=200, command=self.emergency_stop)
        self.em_stop_btn.grid(row=1, column=0, padx=10, pady=5)

        # Load / Unload
        self.unload_btn = ctk.CTkButton(self.btn_frame, text="Unload/Load", width=200, command=self.on_unload_pressed)
        self.unload_btn.grid(row=1, column=1, padx=10, pady=5)

        # Tests
        self.push_test_btn = ctk.CTkButton(self.btn_frame, text="Start Push Test", width=200, command=self.start_push_test)
        self.push_test_btn.grid(row=2, column=0, padx=10, pady=5)
        self.backdrive_test_btn = ctk.CTkButton(self.btn_frame, text="Start Backdrive Test", width=200, command=self.start_backdrive_test)
        self.backdrive_test_btn.grid(row=2, column=1, padx=10, pady=5)

        # Print / Save
        self.print_btn = ctk.CTkButton(self.btn_frame, text="Print Label", width=200, command=self.print_label)
        self.print_btn.grid(row=3, column=0, padx=10, pady=5)
        self.save_data_btn = ctk.CTkButton(self.btn_frame, text="Save Test Data", width=200, command=self.save_data)
        self.save_data_btn.grid(row=3, column=1, padx=10, pady=5)

        # Start reading Arduino
        self.read_arduino()

    # -----------------------------
    # Button Functions
    # -----------------------------
    def emergency_stop(self):
        send_command("STOP")
        messagebox.showwarning("Emergency Stop", "All operations halted!")

    def on_unload_pressed(self):
        confirm = messagebox.askyesno("Confirm", "Are you sure you want to UNLOAD/LOAD the actuator?")
        if not confirm:
            return
        send_command("UNLOAD")

    def start_push_test(self):
        send_command("push test")
        print("Starting push test for Actuator:", self.act_type.get(), "| ID:", self.actuator_id_entry.get())

    def start_backdrive_test(self):
        send_command("backdrive test")
        print("Starting backdrive test for Actuator:", self.act_type.get(), "| ID:", self.actuator_id_entry.get())

    def print_label(self):
        actuator_id = self.actuator_id_entry.get()
        act_type = self.act_type.get()
        push_force = self.push_force_label.cget("text")
        backdrive_force = self.backdrive_force_label.cget("text")

        now = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

        label_text = (
            "------------------------------\n"
            "   LINEAR ACTUATOR TEST DATA\n"
            "------------------------------\n"
            f"Actuator ID: {actuator_id}\n"
            f"Type: {act_type}\n"
            f"Max Push Force: {push_force}\n"
            f"Backdrive Force: {backdrive_force}\n"
            f"Test Time: {now}\n"
            "------------------------------\n"
            "\n\n"
        )

        # Send to Python → Arduino
        send_command("PRINT_LABEL")

        # Windows printing
        self.send_to_printer(label_text)
    

    def save_data(self):
        actuator_id = self.actuator_id_entry.get().strip()
        act_type = self.act_type.get()
        push_force = self.push_force_label.cget("text")
        backdrive_force = self.backdrive_force_label.cget("text")

        if not actuator_id:
            messagebox.showerror("Error", "Please enter Actuator ID before saving.")
            return

        # Save file on desktop
        desktop = os.path.join(os.path.join(os.environ["USERPROFILE"]), "Desktop")
        file_path = os.path.join(desktop, "actuator_test_data.xlsx")

        # If file exists → load it, otherwise create new workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Results"

            # Write the header row ONCE
            ws.append(["Timestamp", "Actuator ID", "Type", "Push Force (N)", "Backdrive Force (N)"])

        # Current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Append a new row of test data
        ws.append([timestamp, actuator_id, act_type, push_force, backdrive_force])

        # Save file
        wb.save(file_path)

        try:
            os.startfile(file_path)  # Windows only
        except Exception as e:
            print("Failed to open Excel file:", e)
    
        print(f"[FILE SAVED] {file_path}")
        messagebox.showinfo("Saved", f"Test data saved to:\n{file_path}")


    # -----------------------------
    # Force Update Functions
    # -----------------------------
    def update_push_force(self, value):
        self.push_force_label.configure(text=f"{value} KG")

    def update_backdrive_force(self, value):
        self.backdrive_force_label.configure(text=f"{value} KG")

    # -----------------------------
    # Arduino Communication
    # -----------------------------
    def read_arduino(self):
        line = arduino.readline().decode(errors='ignore').strip()
        
        if line:
            print("Arduino:", line)

            if line.startswith("PUSH_FORCE_PEAK:"):
                value = line.split(":")[1].strip()
                self.update_push_force(value)
            
            elif line.startswith("BACKDRIVE_FORCE_PEAK: "): 
                value = line.split(":")[1].strip()
                self.update_backdrive_force(value)
            
            elif line.startswith("PUSH_TEST_STATUS:"):
                status = (line.split(":")[1]).strip() 
                color = "green" if status == "PASS" else "red"
                self.push_test_status_label.configure(text=status, text_color=color)

            elif line.startswith("BACKDRIVE_TEST_STATUS:"):
                status = (line.split(":")[1]).strip() 
                color = "green" if status == "PASS" else "red"
                self.backdrive_test_status_label.configure(text=status, text_color=color)

            elif line.startswith("BACKDRIVE TEST COMPLETE"): 
                messagebox.showinfo("Test Update", f"BACKDRIVE TEST COMPLETE")

            elif line.startswith("PUSH TEST STARTED"):  
                messagebox.showinfo("Test Update", f"PUSH TEST STARTED")

            elif line.startswith("PUSH TEST COMPLETE"): 
                messagebox.showinfo("Test Update", f"PUSH TEST COMPLETE")

            elif line.startswith("UNLOAD/LOAD STARTED"): 
                messagebox.showinfo("Process Update", f"UNLOAD/LOAD STARTED")

            elif line.startswith("UNLOAD/LOAD COMPLETE"): 
                messagebox.showinfo("Process Update", f"UNLOAD/LOAD COMPLETE")

      
      
            
            
 
        self.root.after(100, self.read_arduino)  # repeat every 100ms

    def send_to_printer(self, text):
 
        # -------------------------------------------------------
        # 1. Detect a physical hardware printer (not PDF, XPS, Fax)
        # -------------------------------------------------------
        printers = win32print.EnumPrinters(
            win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        )

        hardware_printer = None

        for flags, desc, name, comment in printers:
            lname = name.lower()
            ldesc = desc.lower()

            if ("pdf" in lname or "pdf" in ldesc or
                "xps" in lname or "fax" in lname or
                "onenote" in lname):
                continue  # skip virtual printers

            hardware_printer = name
            break

        if hardware_printer is None:
            print("[ERROR] No physical printer found!")
            messagebox.showerror("Error", "Unable to connect to physical printer.")
            return

        print(f"[INFO] Using printer: {hardware_printer}")
        
        #alert user that print will start shortly as it can be sightly delayed 
        messagebox.showinfo("Print Status", f"Data sent to printer, print will begin shortly")

        # -------------------------------------------------------
        # 2. Create a GDI Device Context for the printer
        # -------------------------------------------------------
        hdc = win32ui.CreateDC()
        hdc.CreatePrinterDC(hardware_printer)

        # -------------------------------------------------------
        # 3. Start the document
        # -------------------------------------------------------
        hdc.StartDoc("Actuator Test Label")
        hdc.StartPage()

        # -------------------------------------------------------
        # 4. Choose the font
        # -------------------------------------------------------
        font = win32ui.CreateFont({
            "name": "Consolas",   # good for fixed-width label text
            "height": 120,         # font size (increase if you want)
            "weight": 1000,
        })
        hdc.SelectObject(font)

        # -------------------------------------------------------
        # 5. Draw multi-line text
        # -------------------------------------------------------
        x = 100  # left margin
        y = 100  # top margin
        line_height = 140

        for line in text.split("\n"):
            hdc.TextOut(x, y, line)
            y += line_height

        # -------------------------------------------------------
        # 6. Finish printing
        # -------------------------------------------------------
        hdc.EndPage()
        hdc.EndDoc()

        print("[INFO] Print job sent successfully")





arduino = connect_arduino()

# ---- Run GUI ----
if __name__ == "__main__":
    root = ctk.CTk()
    app = ActuatorGUI(root)
    root.mainloop()
